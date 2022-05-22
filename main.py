import discord
import replit
from discord.ext import commands
import os
import requests
import json
import music
import random
from random import randint
from random import choice
from replit import db
from keep_alive import keep_alive
import datetime
from datetime import datetime
from openpyxl import load_workbook
# intents = discord.Intents.default()
# intents.typing = False
# intents.presences = False



#import all of the cogs#os.system('pip install pymongo[srv]')
#client = discord.Client()

cogs =[music]

client = commands.Bot(command_prefix="!", intents = discord.Intents.all())


for i in range(len(cogs)):
    cogs[i].setup(client)






sad_words = [
    "depresja", "smutek", "złość", "angry", "oint", "depressing", "dobroć",
    "odpisz", "napisz", "powiedz"
]

starter_encouragements = ["Ale z Ciebie Szef!", "Fajnie", "wtf", "witam"]

matlab = ["matlab"]

memy = ["meme", "memy", "memowanie"]

zajecia = ["zajęcia", "zajecia", "Zajęcia", "Zajecia"]
dzis = ['dzis zajecia' , "dziś zajęcia","dzis zajęcia","dziś zajecia","zajecia dzis", "zajęcia dziś", "zajęcia dzis", "zajecia dziś"]

studia = ["terminarz", "kolokwia", "kolosy", "kolosa", "kolokwium"]

odpowiedz = [
    'https://wutwaw-my.sharepoint.com/:x:/g/personal/01130124_pw_edu_pl/EcivCWFDw89CmseFuSnO4z4BI2NGj_6cstM468jXXl8dCA?e=gYwT7l'
]

db["encouragements"] = []



if "responding" not in db.keys():
    db["responding"] = True


def get_quote():
    response = requests.get("https://zenquotes.io/api/random")
    json_data = json.loads(response.text)
    quote = json_data[0]['q'] + " -" + json_data[0]['a']
    return (quote)


def update_encouragements(encouraging_message):
    if db["responding"]:
        if "encouragements" in db.keys():
            encouragements = db["encouragements"]
            encouragements.append(encouraging_message)
            db["encouragements"] = encouragements
        else:
            db["encouragements"] = [encouraging_message]


def delete_encouragements(index):
    if db["responding"]:
        encouragements = db["encouragements"]
        if len(encouragements) > index:
            del encouragements[index]
            if db["responding"]:
                db["encouragements"] = encouragements





# @client.event
# async def on_ready():
#     print('Jestem gotowy, dawajcie w cs {0.user}'.format(client))










@client.event
async def on_message(message):
  if message.author == client.user:
      return

  msg = message.content


  
  if msg.startswith('inspire'):
      quote = get_quote()
      await message.channel.send(quote)

  if any(word.lower() in msg.lower() for word in studia):
      await message.channel.send("Terminarz kolokwiów: " + str(odpowiedz))

  #if any (word.lower() in msg.lower() for word in zajecia):
  #await message.channel.send( "Jutro nie ma zajęć, gramy gramy GRAMY")

  if db["responding"]:
      options = starter_encouragements
      if "encouragements" in db.keys():
          options = options + list(db['encouragements'])

      if any(word in msg for word in sad_words):
          await message.channel.send(random.choice(options))

  if msg.startswith("new"):
      encouraging_message = msg.split("new ", 1)[1]
      update_encouragements(encouraging_message)
      if db["responding"]:
          await message.channel.send("new encouraging message added.")

  if msg.startswith("del"):
      encouragements = []
      if "encouragements" in db.keys():
          index = int(msg.split("del", 1)[1])
          delete_encouragements(index)
          encouragements = db["encouragements"]
          if db["responding"]:
              await message.channel.send(encouragements)

  if msg.startswith("usuwaj"):
      db["encouragements"] = db["encouragements"].clear()
      db["encouragements"] = []

  if msg.startswith("list"):
      encouragements = []
      if "encouragements" in db.keys():
          encouragements = db["encouragements"]
          if db["responding"]:
              await message.channel.send(encouragements)

  if msg.startswith("responding"):
      value = msg.split("responding ", 1)[1]

      if value.lower() == "true":
          db["responding"] = True
          await message.channel.send("Responding is on.")
      else:
          db["responding"] = False
          await message.channel.send("Responding is off.")

  # if any(word.lower() in msg.lower() for word in memy):
  #     num = randint(1, 8)
  #     try:
  #         await message.channel.send(file=discord.File("{}.jpg".format(num)))
  #     except:
  #         await message.channel.send(file=discord.File("{}.png".format(num)))

  # if any(word.lower() in msg.lower() for word in matlab):
  #     num = randint(9, 9)
  #     try:
  #         await message.channel.send(file=discord.File("{}.jpg".format(num)))
  #     except:
  #         await message.channel.send(file=discord.File("{}.png".format(num)))

  if (message.content.startswith("mazury")):
      await message.channel.send(file=discord.File("9.png"))

  wb = load_workbook('Zeszyt1.xlsx')
  sheet = wb.active 
  if message.author.id in [406436455888715777, 213390565604196352]:
    sheet = wb["Solej"]
  elif message.author.id in [631587778815328256]:
    sheet = wb["Arkusz3"]
  elif message.author.id in [268863457011564547, 373561751926472704]:
    sheet = wb["Tramwaj"]
  elif message.author.id in [74888883061403648]:
    sheet = wb["Ellis"]
  elif message.author.id in [401075326178033685]:
    sheet = wb["Shroomy"]
  elif message.author.id in [191655544463818762]:
    sheet = wb["Piwosz"]
  elif message.author.id in [182554213090328577]:
    sheet = wb["Axword"]
  elif message.author.id in [832003804747530250]:
    sheet = wb["qxv21"]
  elif message.author.id in [213390565604196352]:
    sheet = wb["Marcin"]
  else:
    sheet = wb["Arkusz1"]
  if any(word.lower() in msg.lower() for word in dzis):
      if (datetime.today().weekday()) == 0:
          for row in sheet.iter_rows(min_row=1,
                                     min_col=2,
                                     max_row=20,
                                     max_col=2):
              for cell in row:
                  if (str(cell.value)) == "None":
                      continue
                  else:
                      await message.channel.send(cell.value)
      elif (datetime.today().weekday()) == 1:
          for row in sheet.iter_rows(min_row=1,
                                     min_col=3,
                                     max_row=20,
                                     max_col=3):
              for cell in row:
                  if (str(cell.value)) == "None":
                      continue
                  else:
                      await message.channel.send(cell.value)
      elif (datetime.today().weekday()) == 2:
          for row in sheet.iter_rows(min_row=1,
                                     min_col=4,
                                     max_row=20,
                                     max_col=4):
              for cell in row:
                  if (str(cell.value)) == "None":
                      continue
                  else:
                      await message.channel.send(cell.value)
      elif (datetime.today().weekday()) == 3:
          for row in sheet.iter_rows(min_row=1,
                                     min_col=5,
                                     max_row=20,
                                     max_col=5):
              for cell in row:
                  if (str(cell.value)) == "None":
                      continue
                  else:
                      await message.channel.send(cell.value)
      elif (datetime.today().weekday()) == 4:
          for row in sheet.iter_rows(min_row=1,
                                     min_col=6,
                                     max_row=20,
                                     max_col=6):
              for cell in row:
                  if (str(cell.value)) == "None":
                      continue
                  else:
                      await message.channel.send(cell.value) 
  elif any(word.lower() in msg.lower() for word in zajecia):
    if (datetime.today().weekday()) == 0:
          for row in sheet.iter_rows(min_row=1,
                                     min_col=3,
                                     max_row=20,
                                     max_col=3):
              for cell in row:
                  if (str(cell.value)) == "None":
                      continue
                  else:
                      await message.channel.send(cell.value)
    elif (datetime.today().weekday()) == 1:
          for row in sheet.iter_rows(min_row=1,
                                     min_col=4,
                                     max_row=20,
                                     max_col=4):
              for cell in row:
                  if (str(cell.value)) == "None":
                      continue
                  else:
                      await message.channel.send(cell.value)
    elif (datetime.today().weekday()) == 2:
          for row in sheet.iter_rows(min_row=1,
                                     min_col=5,
                                     max_row=20,
                                     max_col=5):
              for cell in row:
                  if (str(cell.value)) == "None":
                      continue
                  else:
                      await message.channel.send(cell.value)
    elif (datetime.today().weekday()) == 3:
          for row in sheet.iter_rows(min_row=1,
                                     min_col=6,
                                     max_row=20,
                                     max_col=6):
              for cell in row:
                  if (str(cell.value)) == "None":
                      continue
                  else:
                      await message.channel.send(cell.value)
    elif (datetime.today().weekday()) == 6:
          for row in sheet.iter_rows(min_row=1,
                                     min_col=2,
                                     max_row=20,
                                     max_col=2):
              for cell in row:
                  if (str(cell.value)) == "None":
                      continue
                  else:
                      await message.channel.send(cell.value)
    else:
      await message.channel.send("Jutro nie ma zajęć. gramy gramy GRAMY")
  await client.process_commands(message)      
  

@client.event
async def on_command_error(ctx, error):
	if isinstance(error, commands.CommandNotFound):
		await ctx.send('Nie ma takowej komendy')

@client.command(help='Sprawdza ping')
async def ping(ctx):
	await ctx.channel.send(f'pong odbity z prędkością {round(client.latency * 1000)} ms')


@client.command(help='polo')
async def marco(ctx):
	await ctx.channel.send("polo")


@client.command(help='Wypisuje zmiany po najnowszej aktualizacji.')
async def aktualizacja(ctx):
  await ctx.channel.send("Od ostatniej aktualziacji dodano kolejkowanie bota muzycznego, od teraz można zakolejkować kilka piosenek, które odtworzą się jedna po drugiej! Oprócz tego poprawiono wiele pomniejszych błędów. Miłej zabawy ")


@client.command(help='Wysyła mema')
async def memy(ctx):
  szukam = await ctx.channel.send("szukam...")
  for i in range(100):
    url = 'https://meme-api.herokuapp.com/gimme/ProgrammerHumor'
    resp = requests.get(url=url)
    data = resp.json()
    print(data["ups"])
    if data["ups"] > 500:
      break
      # print(data)
      # print(data["url"])

  embed=discord.Embed(title="Memik", url="", description="quality: " + str(data["ups"]), color=0x42cbf5)
  embed.set_image(url=data["url"])
  #msg = data["url"] + " tekst"
  await szukam.delete()
  await ctx.channel.send(embed=embed)
  #await ctx.delete()

print(db["encouragements"])
keep_alive()
my_secret = os.environ['TOKEN']
client.run(my_secret)
